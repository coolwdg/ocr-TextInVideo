import cv2
import os

import openpyxl
from openpyxl.reader.excel import load_workbook
from paddleocr import PaddleOCR
import pandas as pd
import numpy as np
import pytesseract
from natsort import natsorted,ns
from openpyxl import Workbook
from PIL import Image
ocr = PaddleOCR(det=False, lang="ch")
global global_count

def extract_text_from_images(folder_path, output_folder):
    # 创建输出文件夹
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 遍历文件夹中的所有图片文件
    for filename in os.listdir(folder_path):
        if filename.endswith(('.jpg', '.jpeg', '.png')):
            image_path = os.path.join(folder_path, filename)


            # 调用函数进行数字识别并保存结果
            extract_text_from_image(image_path, output_folder)
            os.remove(image_path)


def extract_text_from_image(image_path, output_path):

    image = cv2.imread(image_path)
    x1, y1 = 574, 638
    x2, y2 = 1080, 700
    cropped_image = image[y1:y2, x1:x2]
    global global_count

    result = ocr.ocr(cropped_image,det=False)
    wb = openpyxl.load_workbook(output_path)
    sheet = wb.active
    for item in result:
        row_num = global_count  + 1
        sheet.cell(row=row_num, column=6).value = item[0][0]
        global_count=global_count +1

    wb.save(output_path)





wb = Workbook()
ws = wb.active
path = r"D:\2" #批量截图后的文件夹地址
source_path = r"D:\1" #视频的文件夹地址
dirnames = [f for f in os.listdir(source_path) if os.path.isdir(source_path + '\\' + f)]
print(dirnames)

for  dir_name in dirnames:
    target_path = os.path.abspath(r'D:\2' +'\\' + dir_name)
    if not os.path.exists(target_path):
        # 如果目标路径不存在原文件夹的话就创建
        os.makedirs(target_path)
    filepath = source_path+'\\'+dir_name
    pathDir = os.listdir(filepath)
    pathDir1 = natsorted(pathDir,alg=ns.PATH)
    print(pathDir1)
    for i, filename in enumerate(pathDir1):
        ws.cell(row=i + 1, column=1, value=filename)

        # 保存Excel文件
    wb.save(os.path.join(target_path, f'{dir_name}.xlsx'))


for  dir_name in dirnames:
    target_path = os.path.abspath(r'D:\2' + '\\' + dir_name)
    filepath = source_path + '\\' + dir_name
    pathDir = os.listdir(filepath)
    pathDir1 = natsorted(pathDir, alg=ns.PATH)
    i = 0
    global_count = 0
    for allDir in pathDir1:
        i += 1
        try:
            vc = cv2.VideoCapture(filepath + '/' + allDir)  # 读取视频
            vc.set(cv2.CAP_PROP_POS_MSEC, 0)  # 设置读取位置
            rval, frame = vc.read()  # 读取当前帧，rval用于判断读取是否成功
            if rval:
                video_name = os.path.splitext(allDir)[0]
                cv2.imencode('.jpg', frame)[1].tofile(
                    target_path + '\\' + video_name + '.jpg')  # 将当前帧作为图片保存到 cover_path
            else:
                print("读取失败")
        except Exception as e:
            print(f"获取视频截图失败: {e}")
        folder_path = target_path
        output_folder = target_path  # 保存结果的文件夹路径
        extract_text_from_images(folder_path, target_path + '\\'+f'{dir_name}.xlsx')